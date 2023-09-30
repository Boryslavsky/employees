import csv
import openpyxl
from datetime import datetime

try:
    with open('formatted_employees.csv', mode='r', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=';')
        next(csv_reader)
        employees_data = [row for row in csv_reader]
except FileNotFoundError:
    print("Помилка: Файл CSV не знайдено.")
    exit(1)
except Exception as e:
    print(f"Помилка: {str(e)}")
    exit(1)

def calculate_age(birthdate):
    birthdate = datetime.strptime(birthdate, '%d-%m-%Y')
    today = datetime.today()
    age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
    return age

try:
    wb = openpyxl.Workbook()
    table_all = wb.active
    table_all.title = "all"
    table_all.append(["Прізвище", "Ім’я", "По батькові", "Стать", "Дата народження", "Посада", "Місто проживання", "Адреса проживання", "Телефон", "Email"])

    for i, employee in enumerate(employees_data, start=1):
        table_all.append(employee)

    table_younger_18 = wb.create_sheet(title="younger_18")
    table_younger_18.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

    table_18_45 = wb.create_sheet(title="18-45")
    table_18_45.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

    table_45_70 = wb.create_sheet(title="45-70")
    table_45_70.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

    table_older_70 = wb.create_sheet(title="older_70")
    table_older_70.append(["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"])

    i18=1
    i45=1
    i70=1
    iolder70=1
    for employee in employees_data:
        last_name, first_name, middle_name, gender, birthdate, position, city, address, phone_number, email = employee
        age = calculate_age(birthdate)

        if age < 18:
            row_data = [i18, last_name, first_name, middle_name, birthdate, age]
            table_younger_18.append(row_data)
            i18+=1
        elif 18 <= age <= 45:
            row_data = [i45, last_name, first_name, middle_name, birthdate, age]
            table_18_45.append(row_data)
            i45 += 1
        elif 45 < age <= 70:
            row_data = [i70, last_name, first_name, middle_name, birthdate, age]
            table_45_70.append(row_data)
            i70 += 1
        else:
            row_data = [iolder70, last_name, first_name, middle_name, birthdate, age]
            table_older_70.append(row_data)
            iolder70 += 1

    wb.save('secondFile.xlsx')
    print("XLSX файл успішно створено у файлі 'secondFile.xlsx'.")
except Exception as e:
    print(f"Помилка: {str(e)}")
